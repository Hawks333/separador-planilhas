import streamlit as st
import pandas as pd
import zipfile
from io import BytesIO
from openpyxl import load_workbook, Workbook

st.set_page_config(page_title="Separador de Planilhas com Formatação", layout="centered")

st.title("📊 Separador de Planilha com ou sem Formatação")

st.markdown("""
Envie um arquivo Excel .xlsx com a **primeira linha como cabeçalho** e selecione a coluna para separar os dados.

- Os arquivos separados manterão a **formatação visual original** (cores, bordas, estilos), se possível.
- Se o arquivo estiver corrompido ou com fórmulas problemáticas, você poderá baixar os arquivos **sem formatação**.
""")

uploaded_file = st.file_uploader("📁 Envie seu arquivo .xlsx", type=["xlsx"])

if uploaded_file:
    try:
        # ✅ Lê o conteúdo binário do arquivo
        file_bytes = BytesIO(uploaded_file.read())
        file_bytes.seek(0)

        # ✅ Usa pandas para pré-visualização
        df_preview = pd.read_excel(file_bytes, nrows=5)
        df_preview = df_preview.dropna(axis=1, how="all")
        df_preview = df_preview.loc[:, ~df_preview.columns.str.contains('^Unnamed')]

        coluna_separadora = st.selectbox(
            "Selecione a coluna para separar os arquivos:",
            options=df_preview.columns,
            index=0
        )

        st.success(f"Arquivo carregado com sucesso! Coluna selecionada: **{coluna_separadora}**")
        st.write("Visualização da planilha (5 primeiras linhas):")
        st.write(df_preview)

        # -------------------------------------------------------
        # Botão COM FORMATAÇÃO
        # -------------------------------------------------------
        if st.button("✨ Separar e baixar arquivos com formatação"):
            try:
                file_bytes.seek(0)  # Reposiciona ponteiro
                wb_original = load_workbook(file_bytes)
                ws_original = wb_original.active

                # Leitura segura de cabeçalho com openpyxl
                colunas_openpyxl = []
                for col in range(1, ws_original.max_column + 1):
                    val = ws_original.cell(row=1, column=col).value
                    colunas_openpyxl.append(str(val).strip().lower() if val is not None else "")

                # Normaliza a escolha do usuário (vinda do pandas)
                coluna_normalizada = str(coluna_separadora).strip().lower()

                if coluna_normalizada not in colunas_openpyxl:
                    st.error(f"❌ A coluna **{coluna_separadora}** não foi localizada na planilha.\nVerifique se há espaços invisíveis, acentos ou diferenças de nome.")
                    st.stop()

                idx_coluna_sep = colunas_openpyxl.index(coluna_normalizada) + 1

                dados_por_valor = {}
                for row in ws_original.iter_rows(min_row=2, values_only=False):
                    valor = row[idx_coluna_sep - 1].value
                    if valor:
                        chave = str(valor).strip().lower()
                        dados_por_valor.setdefault(chave, []).append(row)

                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                    for chave, linhas in dados_por_valor.items():
                        wb_novo = Workbook()
                        ws_novo = wb_novo.active

                        # Copia cabeçalhos
                        for col_idx, cell in enumerate(ws_original[1], start=1):
                            if cell.value is None:
                                continue
                            novo_cell = ws_novo.cell(row=1, column=col_idx, value=cell.value)
                            if cell.has_style:
                                novo_cell._style = cell._style

                        # Copia dados
                        for row_idx, row in enumerate(linhas, start=2):
                            for col_idx, cell in enumerate(row, start=1):
                                header = ws_original.cell(row=1, column=col_idx).value
                                if header is None:
                                    continue
                                novo_cell = ws_novo.cell(row=row_idx, column=col_idx, value=cell.value)
                                if cell.has_style:
                                    novo_cell._style = cell._style

                        nome_arquivo = f"{chave}.xlsx".replace("/", "_").replace("\\", "_").replace(":", "-")
                        excel_bytes = BytesIO()
                        wb_novo.save(excel_bytes)
                        excel_bytes.seek(0)
                        zip_file.writestr(nome_arquivo, excel_bytes.read())

                zip_buffer.seek(0)
                st.download_button(
                    label="📥 Baixar arquivos separados com formatação (.zip)",
                    data=zip_buffer,
                    file_name="planilhas_formatadas.zip",
                    mime="application/zip"
                )

            except Exception as e:
                st.error(f"Erro ao tentar manter a formatação: {e}")
                st.info("Você pode tentar a alternativa abaixo para baixar os arquivos sem formatação.")

        # -------------------------------------------------------
        # Botão SEM FORMATAÇÃO
        # -------------------------------------------------------
        if st.button("📁 Separar e baixar arquivos sem formatação (alternativa)"):
            try:
                file_bytes.seek(0)
                df = pd.read_excel(file_bytes)
                df = df.dropna(axis=1, how="all")
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

                df['temp_normalized'] = df[coluna_separadora].astype(str).str.strip().str.lower()

                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                    for valor_normalizado in df['temp_normalized'].dropna().unique():
                        valor_original = df.loc[df['temp_normalized'] == valor_normalizado, coluna_separadora].iloc[0]
                        df_filtrado = df[df['temp_normalized'] == valor_normalizado].copy()
                        df_filtrado = df_filtrado.drop(columns=['temp_normalized'])

                        nome_arquivo = str(valor_original).strip().replace('/', '_').replace('\\', '_').replace(':', '-')
                        excel_bytes = BytesIO()
                        df_filtrado.to_excel(excel_bytes, index=False, engine='openpyxl')
                        excel_bytes.seek(0)
                        zip_file.writestr(f"{nome_arquivo}.xlsx", excel_bytes.read())

                zip_buffer.seek(0)
                st.download_button(
                    label="📥 Baixar arquivos separados (sem formatação)",
                    data=zip_buffer,
                    file_name="planilhas_sem_formatacao.zip",
                    mime="application/zip"
                )

            except Exception as e:
                st.error(f"Erro ao processar versão sem formatação: {e}")

    except Exception as e:
        st.error(f"Erro ao ler o arquivo: {e}")

# Adiciona informação de versão no sidebar
st.sidebar.markdown("---")
st.sidebar.markdown("**Versão 1.0**")
st.sidebar.markdown("*Atualiza automaticamente*")
